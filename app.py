#!/usr/bin/env python3
"""
LISTA DE ELEGIBLES - PROCESO DE ENCARGOS ICBF
Herramienta para generar listas de elegibles según los lineamientos
del Art. 7 del Memorando Radicado No. 20251214000008763 del 16/Jul/2025.
"""

import os
import sys
import uuid
import socket
import hashlib
import logging
import tempfile
import webbrowser
import threading
import multiprocessing
from datetime import datetime

# ================================================================
# FIX: PyInstaller --windowed en Mac redirige stdout/stderr a None
# Esto evita que Flask/Werkzeug crasheen al intentar escribir logs
# ================================================================
if getattr(sys, 'frozen', False):
    if sys.stdout is None or not hasattr(sys.stdout, 'write'):
        sys.stdout = open(os.devnull, 'w')
    if sys.stderr is None or not hasattr(sys.stderr, 'write'):
        sys.stderr = open(os.devnull, 'w')

# Suprimir advertencia de servidor de desarrollo de Flask
log = logging.getLogger('werkzeug')
log.setLevel(logging.ERROR)

from flask import Flask, render_template_string, request, send_file, flash, redirect, url_for, jsonify
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io, zipfile

# PDF
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

app = Flask(__name__)
app.secret_key = 'icbf-encargos-2025'
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024

UPLOAD_FOLDER = tempfile.mkdtemp()
OUTPUT_FOLDER = tempfile.mkdtemp()

# ================================================================
# ESTADO DE JOBS (progreso en tiempo real)
# ================================================================
_jobs = {}  # job_id -> {pct, msg, done, error, stats}

def prog(job_id, pct, msg):
    if job_id and job_id in _jobs:
        _jobs[job_id]['pct'] = pct
        _jobs[job_id]['msg'] = msg

# ================================================================
# HTML TEMPLATES
# ================================================================

HTML_LAYOUT = """
<!DOCTYPE html>
<html lang="es">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Lista de Elegibles - ICBF</title>
    <style>
        * { margin: 0; padding: 0; box-sizing: border-box; }
        body {
            font-family: 'Segoe UI', Arial, sans-serif;
            background: linear-gradient(135deg, #1a3a5c 0%, #2d6a9f 50%, #1a3a5c 100%);
            min-height: 100vh;
            display: flex;
            flex-direction: column;
            align-items: center;
        }
        .header {
            background: white;
            width: 100%;
            padding: 15px 0;
            text-align: center;
            box-shadow: 0 2px 10px rgba(0,0,0,0.2);
        }
        .header h1 {
            color: #1a3a5c;
            font-size: 1.4em;
            margin-bottom: 3px;
        }
        .header p {
            color: #666;
            font-size: 0.85em;
        }
        .container {
            max-width: 800px;
            width: 90%;
            margin: 40px auto;
        }
        .card {
            background: white;
            border-radius: 12px;
            padding: 40px;
            box-shadow: 0 8px 32px rgba(0,0,0,0.2);
            margin-bottom: 20px;
        }
        .card h2 {
            color: #1a3a5c;
            margin-bottom: 20px;
            font-size: 1.3em;
            border-bottom: 2px solid #2d6a9f;
            padding-bottom: 10px;
        }
        .upload-area {
            border: 3px dashed #2d6a9f;
            border-radius: 12px;
            padding: 50px 30px;
            text-align: center;
            cursor: pointer;
            transition: all 0.3s ease;
            background: #f8fbff;
            margin-bottom: 20px;
        }
        .upload-area:hover {
            background: #e8f0fe;
            border-color: #1a3a5c;
        }
        .upload-area.dragover {
            background: #d0e4ff;
            border-color: #0d47a1;
        }
        .upload-icon {
            font-size: 3em;
            color: #2d6a9f;
            margin-bottom: 10px;
        }
        .upload-area p {
            color: #555;
            font-size: 1em;
        }
        .upload-area .hint {
            color: #999;
            font-size: 0.8em;
            margin-top: 8px;
        }
        .btn {
            display: inline-block;
            padding: 14px 40px;
            border: none;
            border-radius: 8px;
            font-size: 1.1em;
            font-weight: bold;
            cursor: pointer;
            transition: all 0.3s ease;
            text-decoration: none;
        }
        .btn-primary {
            background: linear-gradient(135deg, #1a3a5c, #2d6a9f);
            color: white;
        }
        .btn-primary:hover { opacity: 0.9; transform: translateY(-1px); }
        .btn-primary:disabled {
            background: #ccc;
            cursor: not-allowed;
            transform: none;
        }
        .btn-success {
            background: linear-gradient(135deg, #1b8a2e, #27ae60);
            color: white;
        }
        .btn-success:hover { opacity: 0.9; }
        .file-name {
            margin-top: 10px;
            color: #1a3a5c;
            font-weight: bold;
            font-size: 0.95em;
        }
        .criterios {
            background: #f0f7ff;
            border-radius: 8px;
            padding: 20px;
            margin-top: 20px;
        }
        .criterios h3 {
            color: #1a3a5c;
            margin-bottom: 10px;
            font-size: 1em;
        }
        .criterios ol {
            padding-left: 20px;
            color: #444;
            font-size: 0.9em;
            line-height: 1.8;
        }
        .alert {
            padding: 15px 20px;
            border-radius: 8px;
            margin-bottom: 20px;
            font-size: 0.95em;
        }
        .alert-error {
            background: #fde8e8;
            color: #c0392b;
            border-left: 4px solid #c0392b;
        }
        .alert-success {
            background: #e8f8e8;
            color: #1b8a2e;
            border-left: 4px solid #1b8a2e;
        }
        .stats {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(180px, 1fr));
            gap: 15px;
            margin: 20px 0;
        }
        .stat-box {
            background: #f0f7ff;
            border-radius: 8px;
            padding: 15px;
            text-align: center;
        }
        .stat-box .number {
            font-size: 2em;
            font-weight: bold;
            color: #1a3a5c;
        }
        .stat-box .label {
            font-size: 0.8em;
            color: #666;
            margin-top: 3px;
        }
        .loading {
            display: none;
            text-align: center;
            padding: 30px;
        }
        .spinner {
            border: 4px solid #f3f3f3;
            border-top: 4px solid #2d6a9f;
            border-radius: 50%;
            width: 50px;
            height: 50px;
            animation: spin 1s linear infinite;
            margin: 0 auto 15px;
        }
        @keyframes spin {
            0% { transform: rotate(0deg); }
            100% { transform: rotate(360deg); }
        }
        .progress-wrap {
            background: #e0e8f5;
            border-radius: 50px;
            height: 36px;
            overflow: hidden;
            margin: 20px 0 8px;
            box-shadow: inset 0 2px 4px rgba(0,0,0,0.1);
        }
        .progress-bar {
            height: 100%;
            background: linear-gradient(90deg, #1a3a5c, #2d6a9f, #3a8fd1);
            background-size: 200% 100%;
            animation: shimmer 2s linear infinite;
            border-radius: 50px;
            transition: width 0.6s ease;
            display: flex;
            align-items: center;
            justify-content: center;
            min-width: 40px;
        }
        @keyframes shimmer {
            0% { background-position: 200% 0; }
            100% { background-position: -200% 0; }
        }
        .progress-pct {
            color: white;
            font-weight: bold;
            font-size: 0.95em;
            text-shadow: 0 1px 2px rgba(0,0,0,0.3);
        }
        .progress-msg {
            text-align: center;
            color: #555;
            font-size: 0.9em;
            margin-bottom: 20px;
            min-height: 20px;
        }
        .btn-plantilla {
            display: inline-block;
            padding: 8px 18px;
            background: #e8f0fe;
            color: #1a3a5c;
            border: 1px solid #2d6a9f;
            border-radius: 6px;
            font-size: 0.85em;
            font-weight: bold;
            cursor: pointer;
            text-decoration: none;
            transition: all 0.2s;
        }
        .btn-plantilla:hover { background: #d0e4ff; }
        .footer {
            color: rgba(255,255,255,0.6);
            text-align: center;
            padding: 20px;
            font-size: 0.8em;
        }
        form { text-align: center; }
    </style>
</head>
<body>
    <div class="header">
        <h1>Instituto Colombiano de Bienestar Familiar - ICBF</h1>
        <p>Herramienta de Generacion de Lista de Elegibles - Proceso de Encargos</p>
    </div>
    <div class="container">
        {% with messages = get_flashed_messages(with_categories=true) %}
            {% for category, message in messages %}
                <div class="alert alert-{{ category }}">{{ message }}</div>
            {% endfor %}
        {% endwith %}
        {{ content|safe }}
    </div>
    <div class="footer">
        Lineamientos Radicado No. 20251214000008763 del 16 de julio de 2025 | Art. 7 Criterios de Desempate
    </div>
</body>
</html>
"""

HTML_UPLOAD = """
<div class="card">
    <h2>Cargar archivo de manifestaciones</h2>
    <div style="text-align:right; margin-bottom:10px;">
        <a href="/plantilla" class="btn-plantilla">&#11015; Descargar Plantilla Excel</a>
    </div>
    <form method="POST" action="/procesar" enctype="multipart/form-data" id="uploadForm">
        <div class="upload-area" id="dropArea" onclick="document.getElementById('fileInput').click()">
            <div class="upload-icon">&#128206;</div>
            <p>Haz clic aqui o arrastra tu archivo Excel</p>
            <p class="hint">Usa la plantilla descargable. Solo necesitas la hoja "Detalle manifestaciones".</p>
            <div class="file-name" id="fileName"></div>
        </div>
        <input type="file" name="archivo" id="fileInput" accept=".xlsx,.xls" style="display:none"
               onchange="document.getElementById('fileName').textContent = this.files[0] ? this.files[0].name : ''; document.getElementById('btnProcesar').disabled = !this.files[0];">
        <button type="submit" class="btn btn-primary" id="btnProcesar" disabled>
            Generar Lista de Elegibles
        </button>
    </form>
</div>
<div class="card">
    <div class="criterios">
        <h3>Criterios de desempate aplicados (Art. 7 Lineamientos):</h3>
        <ol>
            <li><strong>Grado inmediatamente inferior</strong> - Prioridad al servidor con grado mas cercano al de la vacante (Art. 24, Sec. 5.5)</li>
            <li><strong>Titular de carrera</strong> - Prioridad al servidor que desempene el empleo sobre el cual ostente derechos de carrera</li>
            <li><strong>Evaluacion del Desempeno (EDL)</strong> - Mayor puntaje en la ultima evaluacion definitiva en firme</li>
            <li><strong>Antiguedad en el ICBF</strong> - Mayor antiguedad segun fecha de vinculacion</li>
            <li><strong>Misma dependencia o regional</strong> - Pertenezca a la misma dependencia/regional de la vacante</li>
            <li><strong>Discapacidad</strong> - Servidor que acredite situacion de discapacidad</li>
            <li><strong>Derecho al voto</strong> - <em>Requiere verificacion manual del certificado electoral</em></li>
            <li><strong>Sorteo</strong> - <em>Proceso manual</em></li>
        </ol>
    </div>
</div>
<script>
    const dropArea = document.getElementById('dropArea');
    ['dragenter','dragover'].forEach(e => {
        dropArea.addEventListener(e, ev => { ev.preventDefault(); dropArea.classList.add('dragover'); });
    });
    ['dragleave','drop'].forEach(e => {
        dropArea.addEventListener(e, ev => { ev.preventDefault(); dropArea.classList.remove('dragover'); });
    });
    dropArea.addEventListener('drop', e => {
        const dt = e.dataTransfer;
        const file = dt.files[0];
        if (file) {
            document.getElementById('fileInput').files = dt.files;
            document.getElementById('fileName').textContent = file.name;
            document.getElementById('btnProcesar').disabled = false;
        }
    });
</script>
"""

HTML_RESULTADO = """
<div class="card">
    <h2>&#10003; Lista de Elegibles generada</h2>
    <div class="stats">
        <div class="stat-box">
            <div class="number">{{ empleos }}</div>
            <div class="label">Empleos procesados</div>
        </div>
        <div class="stat-box">
            <div class="number">{{ servidores }}</div>
            <div class="label">Servidores unicos</div>
        </div>
        <div class="stat-box">
            <div class="number">{{ manifestaciones }}</div>
            <div class="label">Candidaturas rankeadas</div>
        </div>
        <div class="stat-box">
            <div class="number">{{ en_carrera }}</div>
            <div class="label">Titulares (Carrera)</div>
        </div>
    </div>
    <div style="text-align:center; margin-top:20px;">
        <a href="/descargar/{{ filename }}" class="btn btn-success" style="font-size:1.1em; padding:14px 40px;">
            &#11015; Descargar Excel Completo
        </a>
    </div>
</div>

<div class="card">
    <h2>&#128196; Descargar PDF por Empleo</h2>
    <p style="color:#666; font-size:0.9em; margin-bottom:15px;">
        Filtra el empleo que necesitas y descarga su PDF individual, o descarga todos de una vez.
    </p>

    <input type="text" id="buscador" placeholder="&#128269; Buscar por nombre, ID o regional..."
        oninput="filtrarEmpleos()"
        style="width:100%; padding:10px 14px; border:1px solid #ccc; border-radius:8px;
               font-size:0.95em; margin-bottom:12px; box-sizing:border-box;">

    <div id="listaEmpleos" style="max-height:320px; overflow-y:auto; border:1px solid #e0e8f5; border-radius:8px;">
        {% for e in empleos_data %}
        <div class="empleo-row" data-buscar="{{ e.empleo_id }} {{ e.cargo|upper }} {{ e.regional|upper }} {{ e.dependencia|upper }}"
             style="display:flex; align-items:center; padding:10px 14px; border-bottom:1px solid #f0f4fa;
                    background:{{ loop.cycle('white','#f8fbff') }}; gap:10px;">
            <div style="flex:1; min-width:0;">
                <div style="font-weight:bold; color:#1a3a5c; font-size:0.9em; white-space:nowrap; overflow:hidden; text-overflow:ellipsis;">
                    ID {{ e.empleo_id }} — {{ e.cargo }}
                </div>
                <div style="color:#888; font-size:0.78em;">
                    Grado {{ e.grado }} | {{ e.dependencia }} | {{ e.regional }} | {{ e.candidatos|length }} candidato(s)
                </div>
            </div>
            <a href="/pdf/{{ sid }}/{{ e.empleo_id }}" target="_blank"
               style="flex-shrink:0; padding:6px 14px; background:linear-gradient(135deg,#1a3a5c,#2d6a9f);
                      color:white; border-radius:6px; font-size:0.82em; font-weight:bold;
                      text-decoration:none; white-space:nowrap;">
                &#11015; PDF
            </a>
        </div>
        {% endfor %}
    </div>

    <div style="text-align:center; margin-top:18px;">
        <a href="/pdfs-zip/{{ sid }}" class="btn btn-primary" style="font-size:1em; padding:12px 32px;">
            &#128230; Descargar TODOS los PDFs (ZIP)
        </a>
    </div>
</div>

<div style="text-align:center; margin-top:5px; margin-bottom:20px;">
    <a href="/" style="color:rgba(255,255,255,0.8); text-decoration:none; font-size:0.9em;">
        &#8592; Procesar otro archivo
    </a>
</div>

<script>
function filtrarEmpleos() {
    const q = document.getElementById('buscador').value.toUpperCase();
    document.querySelectorAll('.empleo-row').forEach(r => {
        r.style.display = r.dataset.buscar.includes(q) ? '' : 'none';
    });
}
</script>
"""

HTML_PROCESANDO = """
<div class="card" style="text-align:center;">
    <h2 style="margin-bottom:6px;">Procesando manifestaciones...</h2>
    <p style="color:#666; font-size:0.9em; margin-bottom:25px;">
        Aplicando criterios de desempate del Art. 7 de los Lineamientos
    </p>

    <div class="progress-wrap">
        <div class="progress-bar" id="barra" style="width:0%">
            <span class="progress-pct" id="pct-txt">0%</span>
        </div>
    </div>
    <div class="progress-msg" id="msg-txt">Iniciando...</div>

    <div style="color:#aaa; font-size:0.8em; margin-top:10px;">
        No cierres esta ventana. El archivo se descargara automaticamente al terminar.
    </div>
</div>
<script>
(function poll() {
    fetch('/progreso/{{ sid }}')
        .then(r => r.json())
        .then(d => {
            document.getElementById('barra').style.width = d.pct + '%';
            document.getElementById('pct-txt').textContent = d.pct + '%';
            document.getElementById('msg-txt').textContent = d.msg;
            if (d.done && !d.error) {
                window.location = '/resultado/{{ sid }}';
            } else if (d.done && d.error) {
                document.getElementById('msg-txt').textContent = 'Error: ' + d.error;
                document.getElementById('msg-txt').style.color = '#c0392b';
            } else {
                setTimeout(poll, 600);
            }
        })
        .catch(() => setTimeout(poll, 1000));
})();
</script>
"""


# ================================================================
# GENERACION DE PDF POR EMPLEO
# ================================================================

AZUL_ICBF   = colors.HexColor('#1F4E79')
AZUL_CLARO  = colors.HexColor('#D6E4F0')
VERDE_FILL  = colors.HexColor('#C6EFCE')
AMARILLO_FILL = colors.HexColor('#FFF2CC')
GRIS_FILL   = colors.HexColor('#F2F2F2')

def generar_pdf_empleo(empleo):
    """Genera el PDF de un empleo con su listado de prelación.
    Usa Paragraph en cada celda para que el texto haga wrap automático.
    """
    buf = io.BytesIO()
    # Márgenes reducidos para aprovechar más el ancho (usable ≈ 27.3 cm)
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=1.2*cm, rightMargin=1.2*cm,
        topMargin=1.3*cm, bottomMargin=1.3*cm
    )

    # --- Estilos de página ---
    st_titulo = ParagraphStyle('titulo',
        fontName='Helvetica-Bold', fontSize=13, textColor=AZUL_ICBF,
        alignment=TA_CENTER, spaceAfter=2)
    st_subtitulo = ParagraphStyle('subtitulo',
        fontName='Helvetica-Bold', fontSize=10, textColor=AZUL_ICBF,
        alignment=TA_CENTER, spaceAfter=3)
    st_info = ParagraphStyle('info',
        fontName='Helvetica', fontSize=8.5, textColor=colors.HexColor('#444444'),
        alignment=TA_CENTER, spaceAfter=5)
    st_norma = ParagraphStyle('norma',
        fontName='Helvetica-Oblique', fontSize=7, textColor=colors.grey,
        alignment=TA_CENTER)

    # --- Estilos de celda (Paragraph wrap automático) ---
    _BASE = dict(fontSize=8, leading=10, wordWrap='CJK')
    st_hdr   = ParagraphStyle('hdr',  fontName='Helvetica-Bold',
                              textColor=colors.white, alignment=TA_CENTER, **_BASE)
    st_norm  = ParagraphStyle('norm', fontName='Helvetica',
                              alignment=TA_LEFT,   **_BASE)
    st_norm_c= ParagraphStyle('nc',   fontName='Helvetica',
                              alignment=TA_CENTER, **_BASE)
    st_bold  = ParagraphStyle('bold', fontName='Helvetica-Bold',
                              alignment=TA_LEFT,   **_BASE)
    st_bold_c= ParagraphStyle('bc',   fontName='Helvetica-Bold',
                              alignment=TA_CENTER, **_BASE)

    def P(txt, center=False, bold=False):
        """Convierte un valor a Paragraph con wrap, negrita y alineación opcionales."""
        s = str(txt) if (txt is not None and str(txt).strip() != '') else '—'
        if bold and center: st = st_bold_c
        elif bold:          st = st_bold
        elif center:        st = st_norm_c
        else:               st = st_norm
        return Paragraph(s, st)

    def PH(txt):
        """Encabezado de columna (blanco, centrado, negrita)."""
        return Paragraph(txt, st_hdr)

    historia = []

    # --- Encabezado del documento ---
    historia.append(Paragraph("INSTITUTO COLOMBIANO DE BIENESTAR FAMILIAR - ICBF", st_titulo))
    historia.append(Paragraph(
        f"Listado de Prelación para Encargo — {empleo['cargo']}", st_subtitulo))
    historia.append(Paragraph(
        f"ID: <b>{empleo['empleo_id']}</b> &nbsp;|&nbsp; "
        f"Cód.: <b>{empleo.get('codigo','') or '—'}</b> &nbsp;|&nbsp; "
        f"Grado: <b>{empleo['grado']}</b> &nbsp;|&nbsp; "
        f"Dependencia: <b>{empleo['dependencia']}</b> &nbsp;|&nbsp; "
        f"Regional: <b>{empleo['regional']}</b>", st_info))
    historia.append(HRFlowable(width="100%", thickness=1.5, color=AZUL_ICBF, spaceAfter=5))

    # --- Encabezados de tabla (13 columnas — eliminado "Dif. Grado" para más espacio) ---
    hdrs = [
        PH('No.'),
        PH('Nombre del\nServidor'),
        PH('Cédula'),
        PH('Cargo\nActual'),
        PH('Dependencia\nActual'),
        PH('Gdo.\nAct.'),
        PH('Situación\n(Carrera/Enc.)'),
        PH('Tipo de\nEvaluación'),
        PH('EDL'),
        PH('Fecha\nIngreso'),
        PH('Antigüedad\n(años)'),
        PH('Discapa-\ncidad'),
        PH('Voto\nElectoral'),
    ]

    # --- Filas de datos ---
    datos_tabla = [hdrs]
    for i, c in enumerate(empleo['candidatos']):
        es_primero = (i == 0)
        datos_tabla.append([
            P(c['ranking'],                                       center=True, bold=es_primero),
            P(c['nombre'],                                                     bold=es_primero),
            P(c['cedula'],                                        center=True, bold=es_primero),
            P(c['cargo_actual'] or '—',                                        bold=es_primero),
            P(c.get('dep_actual') or '—',                                      bold=es_primero),
            P(c['grado_actual'] if c['grado_actual'] != '' else '—',
                                                                  center=True, bold=es_primero),
            P(c['situacion'],                                     center=True, bold=es_primero),
            P(c.get('tipo_edl') or '—',                                        bold=es_primero),
            P(str(c['edl']) if c['edl'] else '—',                center=True, bold=es_primero),
            P(c.get('fecha_ingreso') or '—',                      center=True, bold=es_primero),
            P(str(c['antiguedad']),                               center=True, bold=es_primero),
            P(c['discapacidad'],                                  center=True, bold=es_primero),
            P(c.get('voto') or 'Sin datos',                       center=True, bold=es_primero),
        ])

    # --- Anchos de columna (total ≈ 26.9 cm dentro de 27.3 cm usables) ---
    anchos_col = [
        0.6*cm,   # No.
        4.4*cm,   # Nombre
        2.2*cm,   # Cédula
        3.2*cm,   # Cargo actual
        3.0*cm,   # Dependencia actual
        0.9*cm,   # Grado
        1.9*cm,   # Situación
        3.1*cm,   # Tipo evaluación ("EN PERIODO DE PRUEBA" necesita espacio)
        1.1*cm,   # EDL
        1.8*cm,   # Fecha ingreso
        1.4*cm,   # Antigüedad
        1.4*cm,   # Discapacidad
        1.9*cm,   # Voto electoral
    ]

    tabla = Table(datos_tabla, colWidths=anchos_col, repeatRows=1)
    estilo_tabla = TableStyle([
        # Fondo y texto del encabezado (manejado por ParagraphStyle st_hdr)
        ('BACKGROUND',      (0,0), (-1,0), AZUL_ICBF),
        ('VALIGN',          (0,0), (-1,-1), 'MIDDLE'),
        # Grid
        ('GRID',            (0,0), (-1,-1), 0.4, colors.HexColor('#BBBBBB')),
        # Filas alternas
        ('ROWBACKGROUNDS',  (0,1), (-1,-1), [colors.white, GRIS_FILL]),
        # Padding generoso para legibilidad
        ('TOPPADDING',      (0,0), (-1,-1), 5),
        ('BOTTOMPADDING',   (0,0), (-1,-1), 5),
        ('LEFTPADDING',     (0,0), (-1,-1), 3),
        ('RIGHTPADDING',    (0,0), (-1,-1), 3),
    ])
    # Color especial: #1 verde, #2 amarillo
    if len(datos_tabla) > 1:
        estilo_tabla.add('BACKGROUND', (0,1), (-1,1), VERDE_FILL)
    if len(datos_tabla) > 2:
        estilo_tabla.add('BACKGROUND', (0,2), (-1,2), AMARILLO_FILL)
    tabla.setStyle(estilo_tabla)
    historia.append(tabla)

    # --- Pie de página ---
    historia.append(Spacer(1, 8))
    historia.append(HRFlowable(width="100%", thickness=0.5, color=colors.grey))
    historia.append(Paragraph(
        f"Lineamientos Rad. No. 20251214000008763 del 16/Jul/2025 &nbsp;|&nbsp; Art. 7 Criterios de Desempate &nbsp;|&nbsp; "
        f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')} &nbsp;|&nbsp; "
        f"Total candidatos: {len(empleo['candidatos'])}",
        st_norma))

    doc.build(historia)
    buf.seek(0)
    return buf


# ================================================================
# LOGICA DE PROCESAMIENTO
# ================================================================

def procesar_excel(filepath, job_id=None):
    """Procesa el Excel y genera la Lista de Elegibles."""
    prog(job_id, 5, "Leyendo archivo Excel...")
    xl = pd.ExcelFile(filepath)
    detalle = pd.read_excel(filepath, sheet_name='Detalle manifestaciones')

    # Auto-generar Resumen ID LISTAS si no existe en el archivo
    prog(job_id, 10, "Generando resumen de empleos...")
    if 'Resumen ID LISTAS' in xl.sheet_names:
        resumen_id = pd.read_excel(filepath, sheet_name='Resumen ID LISTAS')
    else:
        # Construir desde Detalle manifestaciones usando columnas del archivo
        col_cargo = 'DENOMINACIÓN' if 'DENOMINACIÓN' in detalle.columns else 'CARGO'
        resumen_id = (
            detalle.dropna(subset=['ID'])
            .sort_values('ID')
            .drop_duplicates(subset=['ID'], keep='first')
            [['ID', col_cargo, 'GRADO', 'DEPENDENCIA', 'REGIONAL']]
            .copy()
        )
        resumen_id = resumen_id.rename(columns={col_cargo: 'CARGO'})
        resumen_id['CODIGO'] = ''
        resumen_id['ESTADO VACANTE'] = 'VACANTE'

    prog(job_id, 15, "Limpiando y validando datos...")
    # Limpiar
    detalle = detalle.dropna(subset=['ID', 'Nombre del Servidor Público', 'Cédula del Servidor Público'])

    # Normalizar SITUACIÓN
    detalle['SITUACION_CLEAN'] = detalle['SITUACIÓN'].str.strip().str.upper()
    detalle['SITUACION_CLEAN'] = detalle['SITUACION_CLEAN'].replace({
        'ENCARGO-VD': 'ENCARGO', 'ENCARGADO': 'ENCARGO'
    })

    # Normalizar DISCAPACIDAD
    detalle['DISCAPACIDAD_CLEAN'] = detalle['DISCAPACIDAD'].str.strip().str.upper()

    prog(job_id, 20, "Calculando criterios de desempate...")
    # Deduplicar: un servidor por empleo
    detalle['ID'] = detalle['ID'].astype(int)
    detalle['completeness'] = detalle[['EDL', 'FECHA INGRESO', 'DISCAPACIDAD']].notna().sum(axis=1)
    detalle = detalle.sort_values('completeness', ascending=False)
    detalle = detalle.drop_duplicates(subset=['ID', 'Cédula del Servidor Público'], keep='first')

    # Calcular criterios
    fecha_ref = datetime.now()
    detalle['FECHA_INGRESO_DT'] = pd.to_datetime(detalle['FECHA INGRESO'], errors='coerce')
    detalle['ANTIGUEDAD_DIAS'] = (fecha_ref - detalle['FECHA_INGRESO_DT']).dt.days
    detalle['ANTIGUEDAD_ANOS'] = round(detalle['ANTIGUEDAD_DIAS'] / 365.25, 1)
    detalle['C1_TITULAR'] = (detalle['SITUACION_CLEAN'] == 'CARRERA').astype(int)
    detalle['C2_EDL'] = pd.to_numeric(detalle['EDL'], errors='coerce').fillna(0)
    detalle['C3_ANTIGUEDAD'] = detalle['ANTIGUEDAD_DIAS'].fillna(0)
    detalle['C5_DISCAPACIDAD'] = (detalle['DISCAPACIDAD_CLEAN'] == 'SI').astype(int)
    detalle['GRADO2_NUM'] = pd.to_numeric(detalle['GRADO2'], errors='coerce').fillna(0)

    # Criterio TIPO DE EDL: DEFINITIVA=2 (arriba), otros=1, EN PERIODO DE PRUEBA=0 (abajo)
    tipo_edl_upper = detalle['TIPO DE EDL'].astype(str).str.strip().str.upper()
    detalle['C_TIPO_EDL'] = tipo_edl_upper.map(
        lambda v: 2 if v == 'DEFINITIVA' else (0 if 'PRUEBA' in v else 1)
    )

    # Criterio vi: Derecho al voto - buscar columna de certificado electoral
    col_voto = next(
        (c for c in detalle.columns
         if any(k in c.upper() for k in ('VOTO', 'ELECTORAL', 'VOTACION', 'SUFRAGIO'))),
        None
    )
    if col_voto:
        detalle['C6_VOTO'] = detalle[col_voto].apply(
            lambda v: 1 if str(v).strip().upper() in ('SI', 'SÍ', '1', 'TRUE', 'X', 'YES') else 0
        )
    else:
        detalle['C6_VOTO'] = 0  # Informacion no disponible aun

    # Criterio vii: Sorteo - orden determinista basado en cedula para reproducibilidad
    def _sorteo_hash(cedula):
        try:
            val = str(int(float(cedula))) if pd.notna(cedula) else '0'
        except Exception:
            val = str(cedula)
        return int(hashlib.md5(val.encode()).hexdigest(), 16) % 100000

    detalle['C7_SORTEO'] = detalle['Cédula del Servidor Público'].apply(_sorteo_hash)

    # Estilos
    hdr_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    hdr_font = Font(bold=True, color='FFFFFF', size=9, name='Arial')
    title_font = Font(bold=True, size=14, name='Arial', color='1F4E79')
    data_font = Font(size=9, name='Arial')
    data_bold = Font(bold=True, size=9, name='Arial')
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    gray_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

    wb = Workbook()
    wb.remove(wb.active)

    # ============================================================
    # HOJA 1: LISTA DE ELEGIBLES
    # ============================================================
    ws1 = wb.create_sheet('LISTA DE ELEGIBLES')

    ws1.merge_cells('A1:V1')
    ws1['A1'] = 'INSTITUTO COLOMBIANO DE BIENESTAR FAMILIAR - ICBF'
    ws1['A1'].font = title_font
    ws1['A1'].alignment = Alignment(horizontal='center')

    ws1.merge_cells('A2:V2')
    ws1['A2'] = 'LISTA DE ELEGIBLES - PROCESO DE ENCARGOS'
    ws1['A2'].font = Font(bold=True, size=13, name='Arial', color='1F4E79')
    ws1['A2'].alignment = Alignment(horizontal='center')

    ws1.merge_cells('A3:V3')
    ws1['A3'] = f'Lineamientos Rad. No. 20251214000008763 del 16/Jul/2025 | Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    ws1['A3'].font = Font(size=9, name='Arial', italic=True)
    ws1['A3'].alignment = Alignment(horizontal='center')

    headers = [
        'No.', 'ID\nEMPLEO', 'CARGO VACANTE', 'CODIGO', 'GRADO\nVACANTE',
        'DEPENDENCIA VACANTE', 'REGIONAL VACANTE', 'ESTADO\nVACANTE',
        'RANKING', 'NOMBRE DEL SERVIDOR', 'CEDULA',
        'CARGO ACTUAL\nSERVIDOR', 'GRADO\nACTUAL', 'DIF.\nGRADO',
        'SITUACION\n(CARRERA/ENCARGO)', 'EDL\n(PUNTAJE)',
        'FECHA DE\nINGRESO', 'ANTIGUEDAD\n(ANOS)',
        'DEPENDENCIA\nACTUAL', 'DISCAPACIDAD',
        'VOTO\nELECTORAL', 'TIPO DE\nEVALUACION'
    ]
    # Posiciones: 17=FECHA INGRESO, 18=ANTIGUEDAD, 19=DEP ACTUAL, 20=DISCAPACIDAD,
    #             21=VOTO, 22=TIPO EVAL

    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=5, column=col, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = center
        cell.border = border
    ws1.row_dimensions[5].height = 40

    widths = [5, 7, 28, 7, 7, 30, 16, 16, 7, 35, 14, 22, 7, 7, 14, 7, 11, 9, 22, 11, 9, 22]
    for i, w in enumerate(widths, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # ============================================================
    # HOJA 2: PRIMER ELEGIBLE POR EMPLEO
    # ============================================================
    ws2 = wb.create_sheet('PRIMER ELEGIBLE POR EMPLEO')
    ws2.merge_cells('A1:L1')
    ws2['A1'] = 'RESUMEN: PRIMER ELEGIBLE POR CADA EMPLEO VACANTE'
    ws2['A1'].font = title_font
    ws2['A1'].alignment = Alignment(horizontal='center')
    ws2.merge_cells('A2:L2')
    ws2['A2'] = 'Servidor con mayor derecho al encargo segun Art. 7 de los Lineamientos'
    ws2['A2'].font = Font(size=10, name='Arial', italic=True)
    ws2['A2'].alignment = Alignment(horizontal='center')

    h2 = ['No.', 'ID', 'CARGO VACANTE', 'GRADO\nVAC.', 'DEPENDENCIA', 'REGIONAL',
          'PRIMER ELEGIBLE', 'CEDULA', 'GRADO\nACTUAL', 'SITUACION', 'EDL', 'ANTIGUEDAD\n(ANOS)', 'TOTAL\nCAND.']
    for col, h in enumerate(h2, 1):
        cell = ws2.cell(row=4, column=col, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = center
        cell.border = border
    ws2.row_dimensions[4].height = 35
    w2 = [5, 7, 28, 7, 30, 16, 35, 14, 7, 12, 7, 9, 7]
    for i, w in enumerate(w2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ============================================================
    # PROCESAR EMPLEOS
    # ============================================================
    current_row = 6
    num_empleo = 0
    total_elegibles = 0
    row2 = 5
    total_titulares = 0
    empleos_data = []   # Para generación de PDFs

    resumen_valido = resumen_id.dropna(subset=['ID'])
    total_empleos = len(resumen_valido)
    prog(job_id, 25, f"Procesando 0 de {total_empleos} empleos...")

    for idx_loop, (_, emp_row) in enumerate(resumen_valido.iterrows()):
        pct_loop = 25 + int((idx_loop / max(total_empleos, 1)) * 60)
        prog(job_id, pct_loop, f"Procesando empleo {idx_loop + 1} de {total_empleos}...")

        empleo_id = int(emp_row['ID'])

        cargo = str(emp_row['CARGO']) if pd.notna(emp_row['CARGO']) else ''
        codigo = emp_row['CODIGO'] if pd.notna(emp_row['CODIGO']) else ''
        grado_emp = emp_row['GRADO'] if pd.notna(emp_row['GRADO']) else ''
        dep_emp = str(emp_row['DEPENDENCIA']) if pd.notna(emp_row['DEPENDENCIA']) else ''
        reg_emp = str(emp_row['REGIONAL']) if pd.notna(emp_row['REGIONAL']) else ''
        estado_vac = str(emp_row['ESTADO VACANTE']).strip() if pd.notna(emp_row['ESTADO VACANTE']) else ''

        candidatos = detalle[detalle['ID'] == empleo_id].copy()
        if len(candidatos) == 0:
            continue

        grado_vac = int(grado_emp) if grado_emp and str(grado_emp).isdigit() else 0

        # Criterio 0: Grado mas cercano a la vacante (inmediatamente inferior)
        candidatos['C0_GRADO'] = candidatos['GRADO2_NUM']

        # Criterio iv: Misma dependencia O regional
        candidatos['C4_MISMA_DEP_REG'] = 0
        if dep_emp:
            mask_dep = candidatos['DEPENDENCIA2'].astype(str).str.strip().str.upper() == dep_emp.strip().upper()
            candidatos.loc[mask_dep, 'C4_MISMA_DEP_REG'] = 1
        if reg_emp:
            mask_reg = candidatos['REGIONAL2'].astype(str).str.strip().str.upper() == reg_emp.strip().upper()
            candidatos.loc[mask_reg & (candidatos['C4_MISMA_DEP_REG'] == 0), 'C4_MISMA_DEP_REG'] = 1

        # Ordenar según criterios Art. 7 + tipo de evaluación
        # TIPO DE EDL es criterio primario: DEFINITIVA arriba, EN PERIODO DE PRUEBA al fondo del listado completo
        # C6_VOTO: Derecho al voto (1=tiene certificado, 0=sin datos/no)
        # C7_SORTEO: desempate final determinista basado en cedula
        candidatos = candidatos.sort_values(
            by=['C_TIPO_EDL', 'C0_GRADO', 'C1_TITULAR', 'C2_EDL', 'C3_ANTIGUEDAD',
                'C4_MISMA_DEP_REG', 'C5_DISCAPACIDAD', 'C6_VOTO', 'C7_SORTEO'],
            ascending=[False, False, False, False, False, False, False, False, False]
        ).reset_index(drop=True)

        num_empleo += 1

        # Escribir en LISTA DE ELEGIBLES
        for rank, (_, cand) in enumerate(candidatos.iterrows(), 1):
            row = current_row

            if cand['C1_TITULAR'] == 1:
                total_titulares += 1

            grado_act = int(cand['GRADO2_NUM']) if cand['GRADO2_NUM'] > 0 else ''
            dif_grado = grado_vac - int(cand['GRADO2_NUM']) if cand['GRADO2_NUM'] > 0 and grado_vac > 0 else ''

            dep_actual = str(cand['DEPENDENCIA2']) if pd.notna(cand.get('DEPENDENCIA2')) else ''
            voto_display = ('SI' if cand['C6_VOTO'] == 1 else 'NO') if col_voto else 'Sin datos'
            tipo_edl_val = str(cand.get('TIPO DE EDL', '')).strip() if pd.notna(cand.get('TIPO DE EDL')) else ''
            fecha_ing_dt = cand.get('FECHA_INGRESO_DT')
            fecha_ing_val = fecha_ing_dt.strftime('%d/%m/%Y') if pd.notna(fecha_ing_dt) else ''

            values = [
                num_empleo if rank == 1 else '',
                empleo_id if rank == 1 else '',
                cargo if rank == 1 else '',
                codigo if rank == 1 else '',
                grado_vac if rank == 1 else '',
                dep_emp if rank == 1 else '',
                reg_emp if rank == 1 else '',
                estado_vac if rank == 1 else '',
                rank,
                cand['Nombre del Servidor Público'],
                str(int(cand['Cédula del Servidor Público'])) if pd.notna(cand['Cédula del Servidor Público']) else '',
                cand['CARGO ACTUAL'] if pd.notna(cand['CARGO ACTUAL']) else '',
                grado_act,
                dif_grado,
                cand['SITUACION_CLEAN'],
                cand['C2_EDL'],
                fecha_ing_val,        # col 17: FECHA DE INGRESO
                cand['ANTIGUEDAD_ANOS'],  # col 18: ANTIGUEDAD ANOS
                dep_actual,           # col 19: DEPENDENCIA ACTUAL
                'SI' if cand['C5_DISCAPACIDAD'] == 1 else 'NO',  # col 20
                voto_display,         # col 21
                tipo_edl_val,         # col 22: TIPO DE EVALUACION
            ]

            for col, val in enumerate(values, 1):
                cell = ws1.cell(row=row, column=col, value=val)
                cell.font = data_bold if rank == 1 else data_font
                cell.border = border
                # center: No, ID, Codigo, Grado vac, Estado vac, Ranking,
                #         Grado act, Dif grado, Situacion, EDL, Fecha ing,
                #         Antiguedad, Discapacidad, Voto
                cell.alignment = center if col in [1,2,4,5,8,9,13,14,15,16,17,18,20,21] else left_align

            ncols = len(headers)
            if rank == 1:
                for col in range(1, ncols + 1):
                    ws1.cell(row=row, column=col).fill = green_fill
            elif rank == 2:
                for col in range(1, ncols + 1):
                    ws1.cell(row=row, column=col).fill = yellow_fill
            elif rank % 2 == 0:
                for col in range(1, ncols + 1):
                    ws1.cell(row=row, column=col).fill = gray_fill

            current_row += 1
            total_elegibles += 1

        # Escribir en PRIMER ELEGIBLE
        primero = candidatos.iloc[0]
        vals2 = [
            num_empleo, empleo_id, cargo, grado_vac, dep_emp, reg_emp,
            primero['Nombre del Servidor Público'],
            str(int(primero['Cédula del Servidor Público'])),
            int(primero['GRADO2_NUM']) if primero['GRADO2_NUM'] > 0 else '',
            primero['SITUACION_CLEAN'],
            primero['C2_EDL'],
            primero['ANTIGUEDAD_ANOS'],
            len(candidatos)
        ]
        for col, val in enumerate(vals2, 1):
            cell = ws2.cell(row=row2, column=col, value=val)
            cell.font = data_font
            cell.border = border
            cell.alignment = center if col in [1,2,4,9,10,11,12,13] else left_align
        if row2 % 2 == 0:
            for col in range(1, len(h2) + 1):
                ws2.cell(row=row2, column=col).fill = gray_fill
        row2 += 1

        # Guardar datos para PDF
        cands_pdf = []
        for rank_pdf, (_, cp) in enumerate(candidatos.iterrows(), 1):
            grado_act_pdf = int(cp['GRADO2_NUM']) if cp['GRADO2_NUM'] > 0 else ''
            dif_pdf = grado_vac - int(cp['GRADO2_NUM']) if cp['GRADO2_NUM'] > 0 and grado_vac > 0 else ''
            voto_pdf = ('SI' if cp['C6_VOTO'] == 1 else 'NO') if col_voto else 'Sin datos'
            tipo_edl_pdf = str(cp.get('TIPO DE EDL', '')).strip() if pd.notna(cp.get('TIPO DE EDL')) else '—'
            fecha_ing_dt_pdf = cp.get('FECHA_INGRESO_DT')
            fecha_ing_pdf = fecha_ing_dt_pdf.strftime('%d/%m/%Y') if pd.notna(fecha_ing_dt_pdf) else '—'
            cands_pdf.append({
                'ranking':      rank_pdf,
                'nombre':       cp['Nombre del Servidor Público'],
                'cedula':       str(int(cp['Cédula del Servidor Público'])) if pd.notna(cp['Cédula del Servidor Público']) else '',
                'cargo_actual': str(cp['CARGO ACTUAL']) if pd.notna(cp['CARGO ACTUAL']) else '',
                'dep_actual':   str(cp['DEPENDENCIA2']) if pd.notna(cp.get('DEPENDENCIA2')) else '',
                'grado_actual': grado_act_pdf,
                'dif_grado':    dif_pdf,
                'situacion':    cp['SITUACION_CLEAN'],
                'tipo_edl':     tipo_edl_pdf,
                'edl':          cp['C2_EDL'] if cp['C2_EDL'] else '',
                'fecha_ingreso': fecha_ing_pdf,
                'antiguedad':   cp['ANTIGUEDAD_ANOS'],
                'discapacidad': 'SI' if cp['C5_DISCAPACIDAD'] == 1 else 'NO',
                'voto':         voto_pdf,
            })
        empleos_data.append({
            'empleo_id':  empleo_id,
            'cargo':      cargo,
            'codigo':     codigo,
            'grado':      grado_vac,
            'dependencia': dep_emp,
            'regional':   reg_emp,
            'estado':     estado_vac,
            'candidatos': cands_pdf,
        })

    # ============================================================
    # HOJA 3: CRITERIOS Y NORMATIVIDAD
    # ============================================================
    ws3 = wb.create_sheet('CRITERIOS Y NORMATIVIDAD')
    ws3.merge_cells('A1:C1')
    ws3['A1'] = 'CRITERIOS DE ORDENAMIENTO Y DESEMPATE'
    ws3['A1'].font = title_font
    ws3['A2'] = 'Memorando Radicado No. 20251214000008763 del 16 de julio de 2025'
    ws3['A2'].font = Font(size=9, name='Arial', italic=True)

    ws3.column_dimensions['A'].width = 12
    ws3.column_dimensions['B'].width = 50
    ws3.column_dimensions['C'].width = 60

    for col, h in enumerate(['PRIORIDAD', 'CRITERIO', 'DESCRIPCION'], 1):
        cell = ws3.cell(row=4, column=col, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.border = border

    criterios_data = [
        ('i.', 'Desempenar el empleo titular (CARRERA)',
         'Sera elegido el servidor que desempene el empleo titular sobre el que ostente derechos de carrera, es decir, que no se encuentre encargado de otro empleo.'),
        ('ii.', 'Evaluacion del Desempeno Laboral (EDL)',
         'Tendra mejor derecho el servidor que registre el mayor puntaje en la ultima Evaluacion del Desempeno Laboral definitiva en firme.'),
        ('iii.', 'Antiguedad en el ICBF',
         'Se tendra en cuenta el servidor con la mayor antiguedad en la entidad, verificando la fecha de vinculacion registrada en la planta de personal.'),
        ('iv.', 'Misma dependencia o regional',
         'Tendra mejor opcion de obtener el encargo el servidor que pertenezca a la misma dependencia o regional en la que se encuentra el empleo a proveer.'),
        ('v.', 'Discapacidad',
         'Se dirimira con el servidor publico que acredite alguna situacion de discapacidad, con el respectivo certificado.'),
        ('vi.', 'Derecho al voto',
         'Se dara aplicacion por analogia a lo dispuesto en el articulo 2 numeral 3 de la Ley 403 de 1997. Los servidores deben acreditarlo con el certificado electoral vigente.'),
        ('vii.', 'Sorteo',
         'De no ser posible el desempate con los criterios anteriores, se decidira mediante sorteo a traves de balotas fisicas o herramientas tecnologicas.'),
    ]
    for i, (num, crit, desc) in enumerate(criterios_data):
        row = 5 + i
        ws3.cell(row=row, column=1, value=num).font = data_bold
        ws3.cell(row=row, column=1).alignment = center
        ws3.cell(row=row, column=2, value=crit).font = data_bold
        ws3.cell(row=row, column=2).alignment = left_align
        ws3.cell(row=row, column=3, value=desc).font = data_font
        ws3.cell(row=row, column=3).alignment = left_align
        for col in range(1, 4):
            ws3.cell(row=row, column=col).border = border
        ws3.row_dimensions[row].height = 45

    # Guardar
    prog(job_id, 90, "Guardando archivo Excel de salida...")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_name = f"LISTA_ELEGIBLES_ICBF_{timestamp}.xlsx"
    output_path = os.path.join(OUTPUT_FOLDER, output_name)
    wb.save(output_path)
    prog(job_id, 100, "¡Completado!")

    stats = {
        'empleos': num_empleo,
        'servidores': detalle['Cédula del Servidor Público'].nunique(),
        'manifestaciones': total_elegibles,
        'en_carrera': total_titulares,
        'filename': output_name,
        'filepath': output_path,
        'empleos_data': empleos_data,
    }
    return stats


# ================================================================
# RUTAS
# ================================================================

@app.route('/')
def index():
    content = render_template_string(HTML_UPLOAD)
    return render_template_string(HTML_LAYOUT, content=content)


@app.route('/procesar', methods=['POST'])
def procesar():
    if 'archivo' not in request.files:
        flash('No se selecciono ningun archivo.', 'error')
        return redirect(url_for('index'))

    archivo = request.files['archivo']
    if archivo.filename == '':
        flash('No se selecciono ningun archivo.', 'error')
        return redirect(url_for('index'))

    if not archivo.filename.lower().endswith(('.xlsx', '.xls')):
        flash('El archivo debe ser un Excel (.xlsx)', 'error')
        return redirect(url_for('index'))

    job_id = uuid.uuid4().hex[:10]
    filepath = os.path.join(UPLOAD_FOLDER, f'{job_id}.xlsx')
    archivo.save(filepath)

    # Verificar que al menos tenga la hoja principal
    try:
        xl = pd.ExcelFile(filepath)
        if 'Detalle manifestaciones' not in xl.sheet_names:
            flash('El archivo debe tener la hoja "Detalle manifestaciones". Descarga la plantilla.', 'error')
            return redirect(url_for('index'))
    except Exception as e:
        flash(f'No se pudo leer el archivo Excel: {str(e)}', 'error')
        return redirect(url_for('index'))

    _jobs[job_id] = {'pct': 0, 'msg': 'Iniciando...', 'done': False, 'error': None, 'stats': None}

    def run_job():
        try:
            stats = procesar_excel(filepath, job_id)
            _jobs[job_id].update({'done': True, 'pct': 100, 'msg': '¡Completado!', 'stats': stats})
        except Exception as e:
            _jobs[job_id].update({'done': True, 'error': str(e), 'msg': 'Error en el procesamiento'})

    t = threading.Thread(target=run_job, daemon=True)
    t.start()

    content = render_template_string(HTML_PROCESANDO, sid=job_id)
    return render_template_string(HTML_LAYOUT, content=content)


@app.route('/progreso/<sid>')
def progreso(sid):
    job = _jobs.get(sid, {'pct': 0, 'msg': 'Esperando...', 'done': False, 'error': None})
    return jsonify(job)


@app.route('/resultado/<sid>')
def resultado(sid):
    job = _jobs.get(sid)
    if not job or not job.get('done'):
        return redirect(url_for('index'))
    if job.get('error'):
        flash(f'Error al procesar: {job["error"]}', 'error')
        return redirect(url_for('index'))
    content = render_template_string(HTML_RESULTADO, sid=sid, **job['stats'])
    return render_template_string(HTML_LAYOUT, content=content)


@app.route('/pdf/<sid>/<int:empleo_id>')
def pdf_empleo(sid, empleo_id):
    job = _jobs.get(sid)
    if not job or not job.get('stats'):
        flash('Sesion expirada. Procesa el archivo nuevamente.', 'error')
        return redirect(url_for('index'))
    empleo = next((e for e in job['stats']['empleos_data'] if e['empleo_id'] == empleo_id), None)
    if not empleo:
        flash(f'Empleo {empleo_id} no encontrado.', 'error')
        return redirect(url_for('index'))
    buf = generar_pdf_empleo(empleo)
    nombre_pdf = f"Prelacion_Encargo_{empleo_id}_{empleo['cargo'][:30].replace(' ','_')}.pdf"
    return send_file(buf, as_attachment=True, download_name=nombre_pdf, mimetype='application/pdf')


@app.route('/pdfs-zip/<sid>')
def pdfs_zip(sid):
    job = _jobs.get(sid)
    if not job or not job.get('stats'):
        flash('Sesion expirada. Procesa el archivo nuevamente.', 'error')
        return redirect(url_for('index'))
    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        for empleo in job['stats']['empleos_data']:
            pdf_buf = generar_pdf_empleo(empleo)
            nombre = f"Prelacion_{empleo['empleo_id']}_{empleo['cargo'][:30].replace(' ','_')}.pdf"
            zf.writestr(nombre, pdf_buf.read())
    zip_buf.seek(0)
    return send_file(zip_buf, as_attachment=True,
                     download_name='PDFs_Lista_Elegibles_ICBF.zip',
                     mimetype='application/zip')


@app.route('/plantilla')
def plantilla():
    """Descarga la plantilla Excel con las columnas correctas."""
    from openpyxl import Workbook as WB
    wb_t = WB()
    ws = wb_t.active
    ws.title = 'Detalle manifestaciones'

    hdr_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    hdr_font = Font(bold=True, color='FFFFFF', size=9, name='Arial')
    border  = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))
    center  = Alignment(horizontal='center', vertical='center', wrap_text=True)

    columnas = [
        'ID', 'DENOMINACIÓN', 'GRADO', 'DEPENDENCIA', 'REGIONAL',
        'Archivo', 'Proceso', 'ID_Registro_Formulario',
        'Hora de inicio', 'Hora de finalización',
        'Nombre del Servidor Público', 'Cédula del Servidor Público',
        'Correo electrónico', 'Relacionado en listado', 'IDs reportados',
        'Observación', 'Motivos observación', 'CARGO ACTUAL', 'GRADO2',
        'DEPENDENCIA2', 'REGIONAL2', 'SITUACIÓN', 'EDL', 'TIPO DE EDL',
        'FECHA INGRESO', 'DISCAPACIDAD', 'ENCARGADO  2025', 'RESOLUCIÓN '
    ]

    anchos = [6,28,6,35,18,30,15,10,18,18,35,14,28,12,40,25,25,28,6,35,18,10,7,12,14,10,10,10]

    for col, (h, w) in enumerate(zip(columnas, anchos), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 30

    tmpfile = os.path.join(OUTPUT_FOLDER, 'PLANTILLA_MANIFESTACIONES.xlsx')
    wb_t.save(tmpfile)
    return send_file(tmpfile, as_attachment=True, download_name='PLANTILLA_MANIFESTACIONES.xlsx')


@app.route('/descargar/<filename>')
def descargar(filename):
    filepath = os.path.join(OUTPUT_FOLDER, filename)
    if os.path.exists(filepath):
        return send_file(filepath, as_attachment=True, download_name=filename)
    flash('Archivo no encontrado. Procesa el Excel nuevamente.', 'error')
    return redirect(url_for('index'))


# ================================================================
# MAIN
# ================================================================

def find_free_port(start=5000, end=5100):
    """Encuentra un puerto disponible en el rango dado."""
    for port in range(start, end):
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
            try:
                s.bind(('127.0.0.1', port))
                return port
            except OSError:
                continue
    return start


def open_browser(port):
    webbrowser.open(f'http://127.0.0.1:{port}')


if __name__ == '__main__':
    multiprocessing.freeze_support()  # Necesario para PyInstaller en Windows

    port = find_free_port()

    print("\n" + "=" * 60)
    print("  LISTA DE ELEGIBLES - PROCESO DE ENCARGOS ICBF")
    print("=" * 60)
    print(f"\n  Aplicacion lista en: http://127.0.0.1:{port}")
    print("  El navegador se abrira en un momento...")
    print("  Para cerrar: presiona Ctrl+C o cierra esta ventana\n")

    threading.Timer(1.5, open_browser, args=[port]).start()
    app.run(host='127.0.0.1', port=port, debug=False, use_reloader=False)
